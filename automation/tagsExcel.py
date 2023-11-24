import boto3
import openpyxl
from contextlib import suppress

region = 'us-east-2'

def data_append(tags, instance_id, resource):
    try:        
        tags_data = []
        key_found = False
        for tag in tags:
            if tag['Key'].startswith('sce:'):
                key_found = True

        if key_found == True:
            for tag in tags:
                if tag['Key'].startswith('sce:'):
                    tags_data.append([resource, instance_id, tag['Key'], tag['Value']])
        else:
            for i in range(6):
                tags_data.append([resource, instance_id, "-", "-"])

        return tags_data
    except Exception as error:
        print(f'excetion in data_append() {str(error)}')

def get_ec2_tags(ec2_client):
    try:
        ec2_instances = ec2_client.describe_instances()['Reservations']
        tags_data = []
        tags = []

        for reservation in ec2_instances:
            for instance in reservation['Instances']:
                instance_id = instance['InstanceId']
                tags = instance.get('Tags', [])

                #tags_data = data_append(tags, instance_id, "EC2")
                key_found = False
                for tag in tags:
                    if tag['Key'].startswith('sce:'):
                        key_found = True

                if key_found == True:
                    for tag in tags:
                        if tag['Key'].startswith('sce:'):
                            tags_data.append(['EC2', instance_id, tag['Key'], tag['Value']])
                else:
                    for i in range(6):
                        tags_data.append(['EC2', instance_id, "-", "-"])

        return tags_data
    except Exception as error:
        print(f'exception in get_ec2_tags() {str(error)}')


def get_rds_tags(rds_client):
    try:
        rds_instances = rds_client.describe_db_instances()['DBInstances']
        tags_data = []

        for instance in rds_instances:            
            db_instance_identifier = instance['DBInstanceIdentifier']
            arn = f"arn:aws:rds:{region}:933919336272:db:{db_instance_identifier}"
            tags = rds_client.list_tags_for_resource(ResourceName=arn).get('TagList', [])
            
            key_found = False
            for tag in tags:
                if tag['Key'].startswith('sce:'):
                    key_found = True

            if key_found == True:
                for tag in tags:
                    if tag['Key'].startswith('sce:'):
                        tags_data.append(['RDS', db_instance_identifier, tag['Key'], tag['Value']])
            else:
                for i in range(6):
                    tags_data.append(['RDS', db_instance_identifier, "-", "-"])                

        return tags_data
    except Exception as error:
        print(f'exception in get_rds_tags() {str(error)}')

def get_elb_tags(elbv2_client):
    try:
        elbs = elbv2_client.describe_load_balancers()['LoadBalancers']
        tags_data = []
        
        for elb in elbs:
            elb_name = elb['LoadBalancerName']
            elb_arn = elb['LoadBalancerArn']
            tags = elbv2_client.describe_tags(ResourceArns=[elb_arn]).get('TagDescriptions', [])
            for tag_description in tags:
                key_found = False
                for tag in tag_description.get('Tags', []):
                    if tag['Key'].startswith('sce:'):
                        key_found = True
                

                if key_found == True:
                    for tag in tag_description.get('Tags', []):
                        if tag['Key'].startswith('sce:'):
                            tags_data.append(['ELB', elb_name, tag['Key'], tag['Value']])
                else:
                    for i in range(6):
                        tags_data.append(['ELB', elb_name, "-", "-"])
                    
        return tags_data        
    except Exception as error:
        print(f'exception in get_elb_tags() {str(error)}')


def get_asg_tags(asg_client):
    try:

        asgs = asg_client.describe_auto_scaling_groups()['AutoScalingGroups']
        tags_data = []

        for asg in asgs:
            asg_name = asg['AutoScalingGroupName']
            tags = asg_client.describe_tags(Filters=[{'Name': 'auto-scaling-group', 'Values': [asg_name]}]).get('Tags', [])
            key_found = False
            for tag in tags:                
                if tag['Key'].startswith('sce:'):
                    key_found = True

            if key_found == True:
                for tag in tags:
                    if tag['Key'].startswith('sce:'):
                        tags_data.append(['ASG', asg_name, tag['Key'], tag['Value']])
            else:
                for i in range(6):
                    tags_data.append(['ASG', asg_name, "-", "-"])

        return tags_data
    except Exception as error:
        print(f'excpetion in get_asg_tags() {str(error)}')

def get_eip_tags(ec2_client):
    try:
        elastic_ips = ec2_client.describe_addresses()['Addresses']
        tags_data = []

        for elastic_ip in elastic_ips:
            allocation_id = elastic_ip['AllocationId']
            tags = ec2_client.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [allocation_id]}]).get('Tags', [])

            key_found = False
            for tag in tags:
                if tag['Key'].startswith('sce:'):
                    key_found = True

            if key_found == True:
                for tag in tags:
                    if tag['Key'].startswith('sce:'):
                        tags_data.append(['EIP', allocation_id, tag['Key'], tag['Value']])
            else:
                for i in range(6):
                    tags_data.append(['EIP', allocation_id, "-", "-"])

        return tags_data
    except Exception as error:
        print(f'excpetion in get_eip_tags() {str(error)}')

def get_s3_buckets():
    try:        
        tags_data = []
        s3 = boto3.client('s3', region_name='us-east-2')
        response = s3.list_buckets()

        buckets = response['Buckets']
        bucket_names = [bucket['Name'] for bucket in buckets]

        for bucket_name in bucket_names:
            tags = s3.get_bucket_tagging(Bucket=bucket_name).get('TagSet', [])
            for tag in tags:
                print(f"{tag['Key']}: {tag['Value']}")
            key_found = False
            for tag in tags:
                if tag['Key'].startswith('sce:'):
                    key_found = True

            if key_found == True:
                for tag in tags:
                    if tag['Key'].startswith('sce:'):
                        tags_data.append(['S3', bucket_name, tag['Key'], tag['Value']])
            else:
                for i in range(6):
                    tags_data.append(['S3', bucket_name, "-", "-"])
            print(tags_data)
        return tags_data

    except Exception as error:
        pass
        #print(f'exception in get_s3_buckets() {str(error)}')

def get_s3_bucket_tags(bucket_name):
    try:
        tags = s3.get_bucket_tagging(Bucket=bucket_name).get('TagSet', [])
        
        for tag in tags:
            print(f"{tag['Key']}: {tag['Value']}")
    except Exception as error:
        print(f'exception in get_s3_buckets_tags() {str(error)}')


def get_s3_tags(s3_client):
    try:
        s3_buckets = s3_client.list_buckets()['Buckets']
        tags_data = []

        for bucket in s3_buckets:
            bucket_name = bucket['Name']
            tags = s3_client.get_bucket_tagging(Bucket=bucket_name).get('TagSet', [])

            key_found = False
            for tag in tags:
                if tag['Key'].startswith('sce:'):
                    key_found = True

            if key_found == True:
                for tag in tags:
                    if tag['Key'].startswith('sce:'):
                        tags_data.append(['S3', bucket_name, tag['Key'], tag['Value']])
            else:
                for i in range(6):
                    tags_data.append(['S3', bucket_name, "-", "-"])

        return tags_data

    except Exception as error:
        print(f"exception in get_s3_tags() {str(error)}")


def get_all_tags():
    try:
        # Create AWS clients for EC2, S3, and RDS
        ec2_client = boto3.client('ec2', region)
        #s3_client = boto3.client('s3', region_name = region, endpoint_url=f'https://s3.us-east-2.amazonaws.com')
        rds_client = boto3.client('rds', region)
        elbv2_client = boto3.client('elbv2', region)
        asg_client = boto3.client('autoscaling', region)


        # Get tags for each resource type
        ec2_tags = get_ec2_tags(ec2_client)        
        #s3_tags = get_new_s3_buckets()
        rds_tags = get_rds_tags(rds_client)
        elb_tags = get_elb_tags(elbv2_client)
        asg_tags = get_asg_tags(asg_client)
        eip_tags = get_eip_tags(ec2_client)


        # Create Excel workbook and sheet
        workbook = openpyxl.Workbook()
        ec2_sheet = workbook.active
        ec2_sheet.title = 'EC2 Tags'
        

        rds_sheet = workbook.create_sheet(title='RDS Tags')
        elb_sheet = workbook.create_sheet(title='LoadBalancer Tags')
        asg_sheet = workbook.create_sheet(title='AutoScalingGroup Tags')
        eip_sheet = workbook.create_sheet(title='ElasticIP Tags')


        # Write header row
        ec2_sheet.append(['Resource Type', 'Resource ID', 'Tag Key', 'Tag Value'])
        rds_sheet.append(['Resource Type', 'Resource ID', 'Tag Key', 'Tag Value'])
        elb_sheet.append(['Resource Type', 'Resource ID', 'Tag Key', 'Tag Value'])
        asg_sheet.append(['Resource Type', 'Resource ID', 'Tag Key', 'Tag Value'])
        eip_sheet.append(['Resource Type', 'Resource ID', 'Tag Key', 'Tag Value'])


        # Append tags for each resource type to the sheet
        for tags_data, sheet in zip([ec2_tags, rds_tags, elb_tags, asg_tags, eip_tags],[ec2_sheet, rds_sheet, elb_sheet, asg_sheet, eip_sheet]):
            for row in tags_data:
                sheet.append(row)

        # Save Excel file
        workbook.save('tags_report.xlsx')


    except Exception as error:
        print(f'exception in get_all_tags() {str(error)}')


def main():
    get_all_tags()

if __name__ == "__main__":
    main()

